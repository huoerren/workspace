import openpyxl
# from matplotlib import pyplot
from statsmodels.tsa import api as smt
# import numpy as np
import os
import math


# def wgn(x, snr):
#     len_x = len(x)
#     Ps = np.sum(np.power(x, 2)) / len_x
#     Pn = Ps / (np.power(10, snr / 10))
#     noise = np.random.randn(len_x) * np.sqrt(Pn)
#     return x + noise


path = r'应变时程'
xlsx_list = os.listdir(path)
try:
    workbook_new = openpyxl.load_workbook('data.xlsx')
except FileNotFoundError:
    workbook_new = openpyxl.Workbook()
worksheet_new = workbook_new.worksheets[0]
num = 0
for num in range(len(xlsx_list)):
    if worksheet_new.max_column == 1 and worksheet_new.cell(1, 1).value is None:
        col = 1
    else:
        col = worksheet_new.max_column + 1
    worksheet_new.cell(1, col, xlsx_list[num])
    # 打开工作簿
    workbook = openpyxl.load_workbook(path + '/' + xlsx_list[num])
    # 获取工作表
    worksheet = workbook.worksheets[0]
    # 读取表中数据
    dis_list = list(worksheet.columns)[1][1:]
    time_list = list(worksheet.columns)[0][1:]
    # 构建时间序列
    time_series = []
    # t = []
    for i in range(len(dis_list)):
        time_series.append(dis_list[i].value)
    # 添加噪声
    # addnoise = wgn(time_series, 10)
    # 获取时间序列长度
    n = len(time_series)
    diff_time = round(time_list[1].value - time_list[0].value, 2)
    # for i in range(-n + 1, n):
    #     t.append(i * diff_time)
    # 计算自相关系数
    acf = smt.stattools.acf(time_series, nlags=n - 1, fft=False, adjusted=True)
    acf = acf.tolist()
    try:
        for i in range(math.ceil(2 / diff_time) + 1):
            worksheet_new.cell(i+2, col, acf[i])
    except IndexError:
        pass
    # acf = acf[len(acf):0:-1] + acf
    # pyplot.plot(t, acf)
    # pyplot.show()
workbook_new.save('data.xlsx')
